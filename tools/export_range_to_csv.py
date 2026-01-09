import csv, json, hashlib, os
from openpyxl import load_workbook

# INPUT
XLSX = "INPUT ANGKUTAN_STOCK_NEW.xlsx"
SHEET = "POSISI TERAKHIR"

# OUTPUT RANGE (tetap yang kamu mau): Y..AH
OUT_MIN_COL = 34  # AH
OUT_MAX_COL = 43  # AQ

# READ RANGE (tambahkan kolom T untuk filter): T..AH
READ_MIN_COL = 20  # T (posisi terakhir)
READ_MAX_COL = 43  # AQ

# ROWS
MIN_ROW = 3        # pastikan ini baris header kamu (kalau header di row 2)
MAX_ROW = 10000

OUT_CSV = "loglist1.csv"
STATE = ".sync_state.json"

def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, str):
        s = v.strip()
        if s == "=":
            return ""
        return s
    return str(v)

def sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def load_state():
    if not os.path.exists(STATE):
        return {}
    with open(STATE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_state(st):
    with open(STATE, "w", encoding="utf-8") as f:
        json.dump(st, f, ensure_ascii=False, indent=2)

def is_invalid_nobtg(x) -> bool:
    if x is None:
        return True
    if isinstance(x, (int, float)):
        return x == 0
    s = str(x).strip()
    return (s == "" or s == "0" or s == "0.0")

def should_skip_posisi(posisi_raw) -> bool:
    """
    Skip kalau posisi terakhir:
    - DKDS (persis)
    - mengandung kata MILIR (MILIR 1-1-2026, MILIR26-11-2025, dll)
    """
    if posisi_raw is None:
        return False
    s = str(posisi_raw).strip().upper()
    if s == "DKDS":
        return True
    if "MILIR" in s:
        return True
    return False

def main():
    # skip kalau Excel tidak berubah
    xhash = sha256_file(XLSX)
    st = load_state()
    if st.get("xlsx_sha256") == xhash:
        print("Excel unchanged; skip export.")
        return

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET}' tidak ditemukan. Ada: {wb.sheetnames}")
    ws = wb[SHEET]

    out_start = OUT_MIN_COL - READ_MIN_COL
    out_end = out_start + (OUT_MAX_COL - OUT_MIN_COL + 1)

    wrote_any = False
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)

        for i, row in enumerate(ws.iter_rows(
            min_row=MIN_ROW,
            max_row=MAX_ROW,
            min_col=READ_MIN_COL,
            max_col=READ_MAX_COL,
            values_only=True
        ), start=MIN_ROW):

            posisi_raw = row[0]  # kolom T
            out_row = row[out_start:out_end]  # Y..AH

            # header wajib ikut (baris pertama di range)
            if i == MIN_ROW:
                w.writerow([cell_str(v) for v in out_row])
                wrote_any = True
                continue

            nobtg_raw = out_row[0]  # kolom Y = noBtg (sesuai output kamu)

            # filter baris:
            if is_invalid_nobtg(nobtg_raw):
                continue

            # skip kalau posisi DKDS / MILIR
            if should_skip_posisi(posisi_raw):
                continue

            w.writerow([cell_str(v) for v in out_row])
            wrote_any = True

    if not wrote_any:
        print("Warning: tidak ada baris data yang lolos filter.")
    st["xlsx_sha256"] = xhash
    save_state(st)
    print(f"Export done -> {OUT_CSV}")

if __name__ == "__main__":
    main()
