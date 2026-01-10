import csv, json, hashlib, os
from datetime import datetime, date
from openpyxl import load_workbook

# ========= INPUT =========
XLSX  = "INPUT_ANGKUTAN_STOCK_NEW.xlsx"
SHEET = "POSISI TERAKHIR"

# baris data dimulai dari row 3 (sesuaikan kalau header kamu di row 3)
MIN_ROW = 3
MAX_ROW = 10000

# Kolom (1-based)
COL_NOBTG  = 2   # B
COL_JENIS  = 8   # H
COL_VOL    = 13  # M
COL_KELAS  = 18  # R
COL_TGL    = 19  # S
COL_POSISI = 20  # T

# ========= OUTPUT =========
OUT_CSV = "stock.csv"
STATE   = ".sync_state_stock.json"

# Stop baca kalau ketemu baris kosong berturut-turut (biar cepat)
MAX_EMPTY_STREAK = 250

def sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def load_state():
    if not os.path.exists(STATE):
        return {}
    with open(STATE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_state(st):
    with open(STATE, "w", encoding="utf-8") as f:
        json.dump(st, f, ensure_ascii=False, indent=2)

def norm_str(v):
    """CATATAN: tidak mengubah '=' jadi kosong (sesuai permintaan)."""
    if v is None:
        return ""
    return str(v).strip()

def is_invalid_nobtg(v) -> bool:
    """
    Patokan baris kosong:
    - noBtg kosong
    - "0", "0.0"
    - "-"
    (TIDAK memasukkan "=")
    """
    if v is None:
        return True
    if isinstance(v, (int, float)):
        return float(v) == 0.0
    s = str(v).strip()
    if s == "":
        return True
    if s == "-":
        return True
    # handle string angka
    if s == "0" or s == "0.0":
        return True
    return False

def safe_float(v):
    try:
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = norm_str(v).replace(",", ".")
        return float(s) if s else 0.0
    except:
        return 0.0

def parse_date(v):
    # Openpyxl bisa return datetime/date kalau cell type date
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v

    s = norm_str(v)
    if not s:
        return None

    # coba beberapa format umum
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except:
            pass
    return None

def should_skip_posisi(posisi_raw) -> bool:
    """
    Skip kayu yang sudah tidak mungkin diangkut:
    - posisi = DKDS (persis)
    - posisi mengandung MILIR
    """
    s = norm_str(posisi_raw).upper()
    if not s:
        return True
    if s == "DKDS":
        return True
    if "MILIR" in s:
        return True
    return False

def main():
    # skip kalau Excel tidak berubah
    xhash = sha256_file(XLSX)
    st = load_state()
    if st.get("xlsx_sha256") == xhash:
        print("Excel unchanged; skip export.")
        #return

    wb = load_workbook(XLSX, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET}' tidak ditemukan. Ada: {wb.sheetnames}")
    ws = wb[SHEET]

    # detail group: (posisi, kelas, jenis) -> btg, vol, last_date
    agg = {}
    last_global = None

    empty_streak = 0
    processed_rows = 0

    for row in ws.iter_rows(min_row=MIN_ROW, max_row=MAX_ROW, values_only=True):
        processed_rows += 1

        nobtg_raw = row[COL_NOBTG - 1]

        # Filter cepat dulu berdasarkan noBtg
        if is_invalid_nobtg(nobtg_raw):
            empty_streak += 1
            if empty_streak >= MAX_EMPTY_STREAK:
                print(f"Stop reading: {MAX_EMPTY_STREAK} baris kosong berturut-turut.")
                break
            continue
        else:
            empty_streak = 0

        # Baru baca kolom lain kalau noBtg valid
        jenis  = norm_str(row[COL_JENIS  - 1])
        vol    = safe_float(row[COL_VOL  - 1])
        kelas  = norm_str(row[COL_KELAS  - 1])
        tgl    = parse_date(row[COL_TGL  - 1])
        posisi = norm_str(row[COL_POSISI - 1])

        if should_skip_posisi(posisi):
            continue

        key = (posisi, kelas, jenis)
        rec = agg.get(key)
        if rec is None:
            rec = {"btg": 0, "vol": 0.0, "last": None}
            agg[key] = rec

        rec["btg"] += 1
        rec["vol"] += vol

        if tgl:
            if rec["last"] is None or tgl > rec["last"]:
                rec["last"] = tgl
            if last_global is None or tgl > last_global:
                last_global = tgl

    # total per posisi + total global
    pos_tot = {}
    glob_btg = 0
    glob_vol = 0.0

    for (posisi, _kelas, _jenis), rec in agg.items():
        pt = pos_tot.get(posisi)
        if pt is None:
            pt = {"btg": 0, "vol": 0.0, "last": None}
            pos_tot[posisi] = pt

        pt["btg"] += rec["btg"]
        pt["vol"] += rec["vol"]
        if rec["last"]:
            if pt["last"] is None or rec["last"] > pt["last"]:
                pt["last"] = rec["last"]

        glob_btg += rec["btg"]
        glob_vol += rec["vol"]

    last_g_str = last_global.strftime("%d-%m-%Y") if last_global else ""

    # tulis CSV
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([
            "posisi", "kelas_diameter", "jenis",
            "btg", "volume_m3",
            "mutasi_terakhir_posisi",
            "mutasi_terakhir_global"
        ])

        # urutkan detail
        items = sorted(agg.items(), key=lambda x: (x[0][0], x[0][1], x[0][2]))

        current_pos = None
        for (posisi, kelas, jenis), rec in items:
            # kalau pindah posisi, tulis TOTAL posisi sebelumnya
            if current_pos is not None and posisi != current_pos:
                pt = pos_tot[current_pos]
                pt_last = pt["last"].strftime("%d-%m-%Y") if pt["last"] else ""
                w.writerow([current_pos, "TOTAL", "", pt["btg"], round(pt["vol"], 3), pt_last, last_g_str])
                w.writerow([])  # pemisah

            current_pos = posisi
            last_pos = rec["last"].strftime("%d-%m-%Y") if rec["last"] else ""
            w.writerow([posisi, kelas, jenis, rec["btg"], round(rec["vol"], 3), last_pos, last_g_str])

        # TOTAL posisi terakhir
        if current_pos is not None:
            pt = pos_tot[current_pos]
            pt_last = pt["last"].strftime("%d-%m-%Y") if pt["last"] else ""
            w.writerow([current_pos, "TOTAL", "", pt["btg"], round(pt["vol"], 3), pt_last, last_g_str])
            w.writerow([])

        # TOTAL GLOBAL
        w.writerow(["GLOBAL", "TOTAL", "", glob_btg, round(glob_vol, 3), last_g_str, last_g_str])

    st["xlsx_sha256"] = xhash
    save_state(st)
    print(f"Export done -> {OUT_CSV}")
    print(f"Processed rows (iter): {processed_rows}, groups: {len(agg)}")

if __name__ == "__main__":
    main()
